import warnings
from io import StringIO
import pandas as pd
import numpy as np
import warnings
from openpyxl import load_workbook
warnings.filterwarnings('ignore')
import re 
from openpyxl import load_workbook
from django.core.files.storage import FileSystemStorage


def read_file(file1, file2, file3, file4):

    d2_df = pd.read_excel(file1, dtype={'D2 (PMI) #': str, 'DIR #':str, 'Reference Documentation (To-Be)\n - Doc#':str }, sheet_name= 'PMI Creation', skiprows= range(0,2))
    d2 = d2_df[['WSWS DMS - USE THIS FOR MD SHEETS', 'D2 (PMI) #', 'DIR #']]
    prt= pd.read_excel(file2, sheet_name= 'PRT Activity Content', skiprows= range(0,6), dtype={'Measurement Box (TRUE/FALSE) (To-Be)':str } )
    pmi= pd.read_excel(file2, sheet_name= 'PMI Generation Info', skiprows= range(0,1))
    print('Data Load Successful...')
    prt = prt.loc[:, ~prt.columns.str.contains('^Unnamed')]
    pmi = pmi.loc[:, ~pmi.columns.str.contains('^Unnamed')]

    print('Merging Sheets into NEW D2 Template...')
    wsws_pmi = pd.merge(pmi, d2, left_on= 'DMS# (To-Be)', right_on= 'WSWS DMS - USE THIS FOR MD SHEETS', how= 'left')
    wsws_pmi['D2 (PMI) #'] =  wsws_pmi['D2 (PMI) #'].fillna('Empty')
    wsws_pmi['DIR #'] =  wsws_pmi['DIR #'].fillna('Empty')
    dms_num = []
    dir_num =[]

    for (i, row) in wsws_pmi.iterrows():
        

        str1 = row['D2 (PMI) #'].replace('Empty', row['DMS# (To-Be)'])
        str2 = row['DIR #'].replace('Empty', row['DIR# (To-Be)'])
        
        dms_num.append(str1)
        dir_num.append(str2)
    
    wsws_pmi['DMS# (To-Be)'] = dms_num
    wsws_pmi['DIR# (To-Be)'] = dir_num
    cols = ['WSWS DMS - USE THIS FOR MD SHEETS', 'D2 (PMI) #', 'DIR #']
    final_pmi = wsws_pmi.drop(cols, axis =1)

    wsws_prt = pd.merge(prt, d2, left_on= 'DMS (To-Be)', right_on= 'WSWS DMS - USE THIS FOR MD SHEETS', how= 'left')
    wsws_prt['D2 (PMI) #'] =  wsws_prt['D2 (PMI) #'].fillna('Empty')
    wsws_prt['DIR #'] =  wsws_prt['DIR #'].fillna('Empty')    

    dms_prt = []
    dir_prt =[]

    for (i, row) in wsws_prt.iterrows():



        str1 = row['D2 (PMI) #'].replace('Empty', row['DMS (To-Be)'])
        str2 = row['DIR #'].replace('Empty', row['DIR to be'])
        


        
        dms_prt.append(str1)
        dir_prt.append(str2)
    

    wsws_prt['DMS (To-Be)'] = dms_prt
    wsws_prt['DIR to be'] = dir_prt
    cols = ['WSWS DMS - USE THIS FOR MD SHEETS', 'D2 (PMI) #', 'DIR #']
    final_prt = wsws_prt.drop(cols, axis =1)

    print('Writing Sheets into WSWS Template...')
    fs = FileSystemStorage()
    f = StringIO('')

    fs.save('d2_prt_sheet.csv', f)
    file_path = fs.path('d2_prt_sheet.csv')
    file_url = fs.url('d2_prt_sheet.csv')

    fs.save('d2_pmi_sheet.csv', f)
    file_path1 = fs.path('d2_pmi_sheet.csv')
    file_url1 = fs.url('d2_pmi_sheet.csv')


    final_prt.to_csv(file_path, index=False)
    final_pmi.to_csv(file_path1, index=False)
    df_s1 = pd.read_csv(file_path, index_col='Unique #', dtype={'DMS (To-Be)': str})
    df_s2 = pd.read_csv(file_path1,  dtype={'DMS# (To-Be)': str})
    df_s2 = df_s2.set_index('DMS# (To-Be)')
    
  
    filename = file4

    def append_df_to_excel(filename, df, sheet_name, startrow,
                           truncate_sheet=False, 
                           **to_excel_kwargs):
        
        

        # ignore [engine] parameter if it was passed
        if 'engine' in to_excel_kwargs:
            to_excel_kwargs.pop('engine')

        writer = pd.ExcelWriter(filename, engine='openpyxl')

        # Python 2.x: define [FileNotFoundError] exception if it doesn't exist 
        try:
            FileNotFoundError
        except NameError:
            FileNotFoundError = IOError


        try:
            # try to open an existing workbook
            writer.book = load_workbook(filename)

            # get the last row in the existing Excel sheet
            # if it was not specified explicitly
            if startrow is None and sheet_name in writer.book.sheetnames:
                startrow = writer.book[sheet_name].max_row

            # truncate sheet
            if truncate_sheet and sheet_name in writer.book.sheetnames:
                # index of [sheet_name] sheet
                idx = writer.book.sheetnames.index(sheet_name)
                # remove [sheet_name]
                writer.book.remove(writer.book.worksheets[idx])
                # create an empty sheet [sheet_name] using old index
                writer.book.create_sheet(sheet_name, idx)

            # copy existing sheets
            writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
        except FileNotFoundError:
            # file does not exist yet, we will create it
            pass

        if startrow is None:
            startrow = 0

        # write out the new sheet
        df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

        # save the workbook
        writer.save()


    # ### Call function and Write to Template

    # In[59]:


    append_df_to_excel(filename, df_s1, sheet_name='PRT Activity Content', startrow=6,
                           truncate_sheet=False)

    append_df_to_excel(filename, df_s2, sheet_name='PMI Generation Info', startrow=1,
                           truncate_sheet=False)
    print('All Tasks Completed.. d2_WSWS Ready for Download...')
    return  []