import warnings
from io import StringIO
import pandas as pd
import numpy as np
import warnings
from openpyxl import load_workbook
warnings.filterwarnings('ignore')
import re 
from openpyxl import load_workbook
from django.core.files.storage import FileSystemStorage


def read_file(file1, file2, file3, file4):
    lmi = pd.read_excel(file1)
    component_library = pd.read_excel( file2, sheet_name= 'Tasks', dtype={'Supresses': str, 'Task order':str, 'Task excluded for:':str })
    tactics_sheet = pd.read_excel(file3, sheet_name= 'Conveyors', skiprows = range(0, 4))
    tactics_sheet_inscope = tactics_sheet[tactics_sheet['TES Project'] !='Not in scope']
    lmi_in_scope = lmi[lmi['In Scope'] == 'Yes']
    lmi_df = lmi_in_scope[['Name', 'Component ID', 'Description', 'Sort Field', 'Default Sequencing']]
    lmi_df.rename(columns={'Description':'Description (from LMI)' , 'Sort Field': 'Conveyor'}, inplace=True)
    component_library_sheet_1_final = component_library.loc[:, ~component_library.columns.str.contains('^Unnamed')]
    wsws = pd.merge(lmi_df, component_library_sheet_1_final, left_on= 'Component ID', right_on= 'Component ID', how= 'inner')
    print('Data Loading Successful')


    tactics_sheet_inscope = tactics_sheet[tactics_sheet['TES Project'] !='Not in scope']
    tactics_sheet_filtered = tactics_sheet_inscope.loc[:,~tactics_sheet_inscope.columns.str.contains('^Unnamed')]
    tactics_sheet_final = tactics_sheet_filtered.loc[:,~tactics_sheet_filtered.columns.duplicated()]
    wsws_1 = tactics_sheet_final.merge(wsws, left_on ='Conveyor', right_on ='Conveyor', how='inner')
    wsws_2 = wsws_1.reset_index(drop= True)
    wsws_2 = wsws_2.astype(str)
    print('Tactics Sheet Merged!!')



    # Frequecny Logic
    freq_in_weeks = []
    print('Running Frequency Logic....')
    for (i, row) in  wsws_2.iterrows():
        str1 = str(row['Frequency']).replace('RPFM', str(row['Motor permalube replace freq.']))
        str2 = str(str1).replace('RPFG', str(row['Gearbox permalube replace freq.']))
        str3 = str(str2).replace('RPFP', str(row['Pulley permalube replace freq.']))
        str4 = ''
        str5 = ''
        str6 = ''
     ##########################################
     ############### Mechanical Strategies:     
        if str(row['Strategy ID']).find("M-VA-ON") != -1:
            str4 = str(str3).replace('L1', str(row['VA Online L1 freq.']))
            str5 = str(str4).replace('L2', 'Yet to be populated')
            str6 = str(str5).replace('L3', 'Yet to be populated')
        
        if str(row['Strategy ID']).find("M-LUBE-OFF") != -1:
            str4 = str(str3).replace('L1', str(row['Lube Offline L1 freq.']))
            str5 = str(str4).replace('L2', str(row['Lube Offline L2 freq.'])) 
            str6 = str(str5).replace('L3', str(row['Lube Offline L3 freq.']))
            
        if str(row['Strategy ID']).find("M-LUBE-ON") != -1:
            str4 = str(str3).replace('L1', str(row['Lube Online L1 freq.']))
            str5 = str(str4).replace('L2', str(row['Lube Online L2 freq.'])) 
            str6 = str(str5).replace('L3', 'Yet to be populated') 
            
        if str(row['Strategy ID']).find("M-INSP-CONV") != -1:
            str4 = str(str3).replace('L1', str(row['Conveyor team inspection L1 freq.']))
            str5 = str(str4).replace('L2', 'Yet to be populated')
            str6 = str(str5).replace('L3', 'Yet to be populated')
            
        if str(row['Strategy ID']).find("M-SKIRT-INSP") != -1:
            str4 = str(str3).replace('L1', str(row['Hard skirt inspection L1 freq.']))
            str5 = str(str4).replace('L2', 'Yet to be populated')
            str6 = str(str5).replace('L3', 'Yet to be populated')
            
        if str(row['Strategy ID']).find("M-SCRAP-INSP") != -1:
            str4 = str(str3).replace('L1', 'Yet to be populated')
            str5 = str(str4).replace('L2', 'Yet to be populated')
            str6 = str(str5).replace('L3', 'Yet to be populated')
            
        if str(row['Strategy ID']).find("M-OIL-SMPL") != -1:
            str4 = str(str3).replace('L1', str(row['Oil sampling freq.']))
            str5 = str(str4).replace('L2', 'Yet to be populated')
            str6 = str(str5).replace('L3', 'Yet to be populated')
            
        if str(row['Strategy ID']).find("M-CLEANERS-INSP") != -1:
            str4 = str(str3).replace('L1', str(row['Belt cleaners inspect & adjust L1 freq.']))
            str5 = str(str4).replace('L2', 'Yet to be populated')
            str6 = str(str5).replace('L3', 'Yet to be populated')  
            
        if str(row['Strategy ID']).find("M-CLEANERS-REPL") != -1:
            str4 = str(str3).replace('L1', str(row['Replace Scrapers L1 freq.']))
            str5 = str(str4).replace('L2', 'Yet to be populated')
            str6 = str(str5).replace('L3', 'Yet to be populated') 
            
        if str(row['Strategy ID']).find("M-CLEANERS-RFRB") != -1:
            str4 = str(str3).replace('L1', str(row['Refurb Scrapers L1 freq.']))
            str5 = str(str4).replace('L2', 'Yet to be populated')
            str6 = str(str5).replace('L3', 'Yet to be populated') 
        
        if str(row['Strategy ID']).find("M-BTT-OFF") != -1:
            str4 = str(str3).replace('L1', str(row['BTT L1 freq.']))
            str5 = str(str4).replace('L2', str(row['BTT L2 freq.']))
            str6 = str(str5).replace('L3', 'Yet to be populated')  
            
        if str(row['Strategy ID']).find("M-CLEANERS-COMM") != -1:
            str4 = str(str3).replace('L1', str(row['Commission Scrapers L1 freq.']))
            str5 = str(str4).replace('L2', 'Yet to be populated')
            str6 = str(str5).replace('L3', 'Yet to be populated')
         
        if str(row['Strategy ID']).find("M-BRAKE-OFF") != -1:
            str4 = str(str3).replace('L1', str(row['Brake Service L1 freq.']))
            str5 = str(str4).replace('L2', str(row['Brake Service L2 freq.']))
            str6 = str(str5).replace('L3', 'Yet to be populated')
            
        if str(row['Strategy ID']).find("M-INSP-FPM") != -1:
            str4 = str(str3).replace('L1', str(row['FPM Conveyor Inspection L1 freq.']))
            str5 = str(str4).replace('L2', str(row['FPM Conveyor Inspection L2 freq.']))
            str6 = str(str5).replace('L3', 'Yet to be populated')
            
        if str(row['Strategy ID']).find("M-INSP-SHTL") != -1:
            str4 = str(str3).replace('L1', str(row['Shuttle inspection L1 freq.']))
            str5 = str(str4).replace('L2', str(row['Shuttle inspection L2 freq.']))
            str6 = str(str5).replace('L3', 'Not yet populated')

        if str(row['Strategy ID']).find("M-SVCE-DRIVE") != -1:
            str4 = str(str3).replace('L1', str(row['Drive Service L1 freq.']))
            str5 = str(str4).replace('L2', str(row['Drive Service L2 freq.']))
            str6 = str(str5).replace('L3', 'Not yet populated')
        
        if str(row['Strategy ID']).find("M-CLEANERS-INSITU") != -1:
            str4 = str(str3).replace('L1', str(row['Refurb Scrapers In-situ L1 freq.']))
            str5 = str(str4).replace('L2', 'Not yet populated')
            str6 = str(str5).replace('L3', 'Not yet populated')
        
        if str(row['Strategy ID']).find("M-VA-SEV2") != -1:
            str4 = str(str3).replace('L1', str(row['VA Severity 2 freq.']))
            str5 = str(str4).replace('L2', 'Not yet populated')
            str6 = str(str5).replace('L3', 'Not yet populated')   
        ##########################################
        ############### Electrical Strategies:          
        if str(row['Strategy ID']).find("E-MOTR-INSP") != -1:
            str4 = str(str3).replace('L1', str(row['Motor Inspection L1 freq.']))
            str5 = str(str4).replace('L2', 'Yet to be populated')
            str6 = str(str5).replace('L3', 'Yet to be populated')      
        
        if str(row['Strategy ID']).find("E-MOTR-SVCE") != -1:
            str4 = str(str3).replace('L1', str(row['Motor Service L1 freq.']))
            str5 = str(str4).replace('L2', str(row['Motor Service L2 freq.']))
            str6 = str(str5).replace('L3', str(row['Motor Service L3 freq.']))
            
        if str(row['Strategy ID']).find("E-MOTR-SVCE-SPEC") != -1:
            str4 = str(str3).replace('L1', str(row['Motor Service - Contractor L1 freq.']))
            str5 = str(str4).replace('L2', 'Yet to be populated')
            str6 = str(str5).replace('L3', 'Yet to be populated')
               
        if str(row['Strategy ID']).find("E-MOTR-TEST") != -1:
            str4 = str(str3).replace('L1', str(row['Motor TEST - L1 freq.']))
            str5 = str(str4).replace('L2', 'Yet to be populated')
            str6 = str(str5).replace('L3', 'Yet to be populated')       
            
        if str(row['Strategy ID']).find("E-CONV-SVCE") != -1:
            str4 = str(str3).replace('L1', str(row['E-CONV-SVCE - L1 freq.']))
            str5 = str(str4).replace('L2', str(row['E-CONV-SVCE - L2 freq.']))
            str6 = str(str5).replace('L3', 'Yet to be populated')
            
        if str(row['Strategy ID']).find("E-STAT-CONV") != -1:
            str4 = str(str3).replace('L1', str(row['E-STAT-CONV - L1 freq']))
            str5 = str(str4).replace('L2', str(row['E-STAT-CONV - L2 freq']))
            str6 = str(str5).replace('L3', 'Yet to be populated')
                 
        if str(row['Strategy ID']).find("E-STAT-SIREN") != -1:
            str4 = str(str3).replace('L1', str(row['E-STAT-Siren - L1 freq']))
            str5 = str(str4).replace('L2', 'Yet to be populated')
            str6 = str(str5).replace('L3', 'Yet to be populated')
        
        if str(row['Strategy ID']).find("E-MAGNET-INSP") != -1:
            str4 = str(str3).replace('L1', str(row['Magnet Insp - L1 freq']))
            str5 = str(str4).replace('L2', str(row['Magnet Insp - L2 freq']))
            str6 = str(str5).replace('L3', 'Yet to be populated')
            
        if str(row['Strategy ID']).find("E-MAGNET-SVCE") != -1:
            str4 = str(str3).replace('L1', str(row['Magnet Svce - L1 freq']))
            str5 = str(str4).replace('L2', str(row['Magnet Svce - L2 freq']))
            str6 = str(str5).replace('L3', 'Yet to be populated')
            
        if str(row['Strategy ID']).find("E-TMD-SVCE") != -1:
            str4 = str(str3).replace('L1', str(row['TMD Service - L1 freq']))
            str5 = str(str4).replace('L2', str(row['TMD Service - L2 freq']))
            str6 = str(str5).replace('L3', 'Yet to be populated')
            
        if str(row['Strategy ID']).find("E-WEIGHT-SVCE") != -1:
            str4 = str(str3).replace('L1', str(row['Weightometer Svce - L1 freq']))
            str5 = str(str4).replace('L2', str(row['Weightometer Svce - L2 freq']))
            str6 = str(str5).replace('L3', str(row['Weightometer Svce - L3 freq']))
        
        if str(row['Strategy ID']).find("E-CONV-GEN-INSP") != -1:
            str4 = str(str3).replace('L1', str(row['Conveyor General Inspection - L1 freq']))
            str5 = str(str4).replace('L2', str(row['Conveyor General Inspection - L2 freq']))
            str6 = str(str5).replace('L3', 'Yet to be populated')
            
        if str(row['Strategy ID']).find("E-MOISTURE-SVCE") != -1:
            str4 = str(str3).replace('L1', str(row['Moisture Analyser - L1 freq']))
            str5 = str(str4).replace('L2', str(row['Moisture Analyser - L2 freq']))
            str6 = str(str5).replace('L3', str(row['Moisture Analyser - L3 freq']))
        
        if str(row['Strategy ID']).find("E-COOLING-INSP") != -1:
            str4 = str(str3).replace('L1', str(row['Cooling System Inspection - L1']))
            str5 = str(str4).replace('L2', 'Yet to be populated')
            str6 = str(str5).replace('L3', 'Yet to be populated')


        if str(row['Strategy ID']).find("E-ENCD-REPL") != -1:
            str4 = str(str3).replace('L1', str(row['Shuttle Encoder Replacement - L1 freq']))
            str5 = str(str4).replace('L2', 'Yet to be populated')
            str6 = str(str5).replace('L3', 'Yet to be populated')
    #    
        freq_in_weeks.append(str6)
        

    wsws_2['Frequency_In_Weeks'] = freq_in_weeks

    wsws_2['Frequency_In_Weeks'] = wsws_2['Frequency_In_Weeks'].str.replace('nan', '0W').replace('', '0W')
    wsws_2 =  wsws_2[wsws_2['Frequency_In_Weeks'] != 'Strategy N/A']
    print('Frequency Logic Completed!!!')


    ## Acceptable Limit
    print('Running Acceptable Limit...')
    #wsws_2['new_replace_perma'] = wsws_2['Pulley permalube replace freq.'].str.replace('Strategy N/A', '0W')
    acceptable_limit = []

    for (i, row) in  wsws_2.iterrows():
        str1 = str(row['Acceptable limit']).replace('{TS - width}', str(row['Belt width']))
        str2 = str(str1).replace('{TS - proxy setup}', str(row['Proxy setup']))
        str3 = str(str2).replace('"{TS - RPFM} + 2W"',
        (str(([x for x in (list(map(int, (re.findall('\d+', row['Motor permalube replace freq.'].replace('nan', '0').replace('Strategy N/A', '0W'))))))][0]) + 
             [x for x in (list(map(int, (re.findall('\d+', '"{TS - RPFM} + 2W"')))))][0]) + " Weeks"))
        
        str4 = str(str3).replace('"{TS - RPFG} + 2W"',
        (str(([x for x in (list(map(int, (re.findall('\d+', row['Gearbox permalube replace freq.'].replace('nan', '0').replace('Strategy N/A', '0W'))))))][0]) + 
              [x for x in (list(map(int, (re.findall('\d+', '"{TS - RPFG} + 2W"')))))][0]) + " Weeks"))
        
        str5 = str(str4).replace('"{TS - RPFP}+2W"',
        (str(([x for x in (list(map(int, (re.findall('\d+', row['Pulley permalube replace freq.'].replace('nan', '0').replace('Strategy N/A', '0W'))))))][0]) + 
              [x for x in (list(map(int, (re.findall('\d+', '"{TS - RPFP}+2W"')))))][0]) + " Weeks"))
        
        str6 = str(str5).replace('"{TS - M-LUBE-OFF_L1} + 2W"',
        (str(([x for x in (list(map(int, (re.findall('\d+', row['Lube Offline L1 freq.'].replace('nan', '0').replace('Strategy N/A', '0W'))))))][0]) + 
              [x for x in (list(map(int, (re.findall('\d+', '"{TS - RPFP}+2W"')))))][0]) + " Weeks"))
        
        
        str7 = str(str6).replace('{TS - Lanyard Rod Measurement}', 
               str(row['Lanyard Rod Measurement '])).replace('{TS - Skirt clearance limit }',
               str(row['Skirt clearance limit'])).replace('{TS - Conveyor}',
               str(row['Conveyor'])).replace('{TS - Primary scraper model}',
               str(row['Primary scraper model'])).replace('{TS - Primary Scraper Tensioning Arrangement}',
               str(row['Primary Scraper Tensioning Arrangement'])).replace('{TS - Secondary scraper model}',
               str(row['Secondary Scraper Model'])).replace('{TS - Secondary Scraper Tensioning Arrangement}',
               str(row['Secondary Scraper Tensioning Arrangement'])).replace('{TS - Tertiary scraper model}',
               str(row['Tertiary Scraper '])).replace('{TS - Tertiary Scraper Tensioning Arrangement}',
               str(row['Tertiary Scraper Tensioning Arrangement'])).replace('{TS - Magnet Cold Amps}',
               str(row['Magnet Cold Amps'])).replace('{TS - Magnet Hot Amps}',
               str(row['Magnet Hot Amps'])).replace('{TS - Power Unit 1}',
               str(row['Power Unit 1'])).replace('{TS - Caliper 1}',
               str(row['Caliper 1 '])).replace('{TS - Caliper 2}',
               str(row['Caliper 2'])).replace('{TS - SP1}',
               str(row['SP1'])).replace('{TS - SP2}',
               str(row['SP2'])).replace('{TS - RP1}',
               str(row['RP1'])).replace('{TS - RP2}',
               str(row['RP2'])).replace('{TS - Shuttle Power Unit}',
               str(row['Shuttle Power Unit'])).replace('{TS - Shuttle Caliper}',
               str(row['Shuttle Caliper'])).replace('{TS - Shuttle SP1}',
               str(row['Shuttle SP1'])).replace('{TS - Shuttle SP2}',
               str(row['Shuttle SP2'])).replace('{TS - Shuttle RP2}',
               str(row['Shuttle RP2'])).replace('{TS - Shuttle RP1}',
               str(row['Shuttle RP1'])).replace('{TS - Relief Pressure}',
               str(row['Relief Pressure'])).replace('{TS - Accumulator Pre-charge}',
               str(row['Accumulator Pre-charge'])).replace('{TS - Shuttle Relief Pressure}',
               str(row['Shuttle Relief Pressure'])).replace('{TS - Shuttle Accumulator Pre-charge}',
               str(row['Shuttle Accumulator Pre-charge'])).replace('{TS - Magnet Winding Resistance}',
               str(row['Magnet Winding Resistance'])).replace('{TS - Shuttle Wheel Minimum Diameter}',
               str(row['Shuttle Wheel Minimum Diameter'])).replace('{TS - Shuttle Wheel Minimum Flange Thickness}',
               str(row['Shuttle Wheel Minimum Flange Thickness'])).replace('{Shuttle Wheel Flange Wear Limit}',
               str(row['Shuttle Wheel Flange Wear Limit'])).replace('{Shuttle Wheel Design Flange Thickness}.',
               str(row['Shuttle Wheel Design Flange Thickness'])).replace('{TS - Skirt clearance limit}',
               str(row['Skirt clearance limit'])).replace('{TS - Organic Pad Max Wear}',
               str(row['Organic Pad Max Wear'])).replace('{TS - Organic Pad Replace Thickness}',
               str(row['Organic Pad Replace Thickness'])).replace('{TS - Sintered Pad Max Wear}',
               str(row['Sintered Pad Max Wear'])).replace('{TS - Sintered Pad Replace Thickness}',
               str(row['Sintered Pad Replace Thickness'])).replace('{TS - Shuttle Organic Pad Max Wear}',
               str(row['Shuttle Organic Pad Max Wear'])).replace('{TS - Shuttle Organic Pad Replace Thickness}',
               str(row['Shuttle Organic Pad Replace Thickness'])).replace('{TS - Shuttle Sintered Pad Max Wear}',
               str(row['Shuttle Sintered Pad Max Wear'])).replace('{TS - Shuttle Sintered Pad Replace Thickness}',
               str(row['Shuttle Sintered Pad Replace Thickness'])).replace('{TS - Fluid Coupling Model}',
               str(row['Fluid Coupling Model'])).replace('{TS - Fluid Coupling Oil Qty}',
               str(row['Fluid Coupling Oil Qty'])).replace('{TS - Fluid Coupling Oil Type}',
               str(row['Fluid Coupling Oil Type'])).replace('{TS - Fluid Coupling Fusible Plug Colour}',
               str(row['Fluid Coupling Fusible Plug Colour'])).replace('{TS - Fluid Coupling Filler Plug Torque}',
               str(row['Fluid Coupling Filler Plug Torque'])).replace('{TS - Fluid Coupling Fusible Plug Torque}',
               str(row['Fluid Coupling Fusible Plug Torque'])).replace('{TS - Fluid Coupling Screw Plug Torque}',
               str(row['Fluid Coupling Screw Plug Torque']))

        acceptable_limit.append(str7)

    
    #### `Corrective Action`
    c_action = []

    for (i, row) in wsws_2.iterrows():

        str1 = str(row['Corrective action']).replace('{TS - Accumulator Pre-charge}', 
               str(row['Accumulator Pre-charge'])).replace('{TS - Shuttle Accumulator Pre-charge}',
               str(row['Shuttle Accumulator Pre-charge']))
    
        c_action.append(str1)
    
    wsws_2['Corrective action'] = c_action

    wsws_2['Acceptable limit'] = acceptable_limit
    print('Acceptable Limit  and  Corrective Action Completed!!!')


    print('Running Task Description Logic....')


    task = []
    for (i, row) in  wsws_2.iterrows():
        
        str1 = str(row['Task description']).replace('{TS - BTT WIN}',
               str(row['BTT WIN'])).replace('{TS - Conveyor}',
               str(row['Conveyor'])).replace('{TS - Primary Replace WIN No.}',
               str(row['Primary Replace WIN No.'])).replace('{TS - Primary Replace WIN Title}',
               str(row['Primary Replace WIN Title'])).replace('{TS - Secondary Replace WIN No.}',
               str(row['Secondary Replace WIN No.'])).replace('{TS - Secondary Replace WIN Title}',
               str(row['Secondary Replace WIN Title'])).replace('{TS - Tertiary Replace WIN No.}',
               str(row['Tertiary Replace WIN No.'])).replace('{TS - Tertiary Replace WIN Title}',
               str(row['Tertiary Replace WIN Title'])).replace('{TS - Primary Refurb WIN No.}',
               str(row['Primary Refurb WIN No.'])).replace('{TS - Primary Refurb WIN Title}',
               str(row['Primary Refurb WIN Title'])).replace('{TS - Secondary Refurb WIN No.}',
               str(row['Secondary Refurb WIN No.'])).replace('{TS - Secondary Refurb WIN Title}',
               str(row['Secondary Refurb WIN Title'])).replace('{TS - Tertiary Refurb WIN No.}',
               str(row['Tertiary Refurb WIN No.'])).replace('{TS - Tertiary Refurb WIN Title}',
               str(row['Tertiary Refurb WIN Title'])).replace('{TS - Filter PMAC Codes}', 
               str(row['Filter PMAC Codes'])).replace('{TS - Power Unit 1}',
               str(row['Power Unit 1'])).replace('{TS - Caliper 1}',
               str(row['Caliper 1 '])).replace('{TS - Caliper 2}',
               str(row['Caliper 2'])).replace('{TS - SP1}',
               str(row['SP1'])).replace('{TS - SP2}',
               str(row['SP2'])).replace('{TS - RP1}',
               str(row['RP1'])).replace('{TS - RP2}',
               str(row['RP2'])).replace('{TS - Shuttle Power Unit}',
               str(row['Shuttle Power Unit'])).replace('{TS - Shuttle Caliper}',
               str(row['Shuttle Caliper'])).replace('{TS - Shuttle SP1}',
               str(row['Shuttle SP1'])).replace('{TS - Shuttle SP2}',
               str(row['Shuttle SP2'])).replace('{TS - Shuttle RP2}',
               str(row['Shuttle RP2'])).replace('{TS - Shuttle RP1}',
               str(row['Shuttle RP1'])).replace('{Shuttle Wheel Design Flange Thickness}',
               str(row['Shuttle Wheel Design Flange Thickness'])).replace('{TS - Relief Pressure}',
               str(row['Relief Pressure'])).replace('{TS - Shuttle Relief Pressure}',
               str(row['Shuttle Relief Pressure'])).replace('{TS - Fluid Coupling Oil Qty}',
               str(row['Fluid Coupling Oil Qty'])).replace('{TS - Fluid Coupling Oil Type}',
               str(row['Fluid Coupling Oil Type'])).replace('{TS - Fluid Coupling Oil Qty}',
               str(row['Fluid Coupling Oil Qty']))
        
        task.append(str1)
        
    wsws_2['Task description'] = task
    print('Task Description Completd....')



    print('Spliting the Data for Exclusive and Excluded Logic....')

    #Splitting the Data
    pex = []

    for (i, row) in  wsws_2.iterrows():
        
        str1 = str(row['Task excluded for:'])
        
        if str(row['Task excluded for:']) != 'nan':
            
            str1 = 'Excluded: ' + str(row['Task excluded for:'])
            
        pex.append(str1)
        
    wsws_2['new_task_excluded_for'] = pex



    exclu = []
    for (i, row) in  wsws_2.iterrows():
        
        str1 = str(row['Task exclusive to:'])
        
        if str(row['Task exclusive to:']) != 'nan':
            
            str1 = 'Exclusive: ' + str(row['Task exclusive to:'])
            
        exclu .append(str1)
        
    wsws_2['new_task_exclusive_to'] = exclu

    wsws_2['split_rule'] = wsws_2.apply(lambda x: x['new_task_exclusive_to'] 
                                        if x['new_task_excluded_for'] == 'nan' 
                                        else x['new_task_excluded_for'], axis=1)

    pass_data = wsws_2[wsws_2['split_rule'] == 'nan']
    logic_data = wsws_2[(wsws_2['split_rule'] != 'nan')]
    exclud_logic_data = logic_data[logic_data['split_rule'].str.startswith('Excluded:')]
    exclusiv_logic_data= logic_data[logic_data['split_rule'].str.startswith('Exclusive:')]

    print('Data Split Completed!!!')
    

    print('Running Task Excluded Logic....')
    exclud = []

    for (i, row) in  exclud_logic_data.iterrows():
        
        shuttle_numbers = str(row['Shuttle pulleys']).split(',')
        pulley_number =  re.findall(r'^Pulley (\d+)', str(row['Description (from LMI)']))
        
        str1 = 'Excluded_Rule_Keep'
        
        if ((str(row['split_rule']).find('Excluded: Pulley numbers (from LMI list)') != -1 )
         and (any(number in pulley_number for number in shuttle_numbers) == True)):
            
            str1 = 'Pulley_Excluded_Rule_Passed'
            
            
        elif (str(row['split_rule']) == 'Excluded: {TS - Drive Motor/s Included in General Inspection} = Y'
             and (row['Drive Motor/s Included in General Inspection'] == 'Y')
             and (row['Strategy ID'] == 'E-MOTR-INSP')
             and ( (row['Component ID']=='LV_motor_no-RTDs') |
                   (row['Component ID']=='LV_motor')  |
                   (row['Component ID']=='HV_scim_motor'))
            ):
             str1 = 'Drive_Motors_Excluded_Rule_Passed'
                   
        elif (str(row['split_rule'])== 'Excluded: {TS - Filters Replaced every 20Ws} = No Filter Used' 
             and (row['Filters Replaced every 20Ws'] == 'No Filter Used' )
             and (row['Strategy ID'] == 'E-MOTR-SVCE') 
             and (row['Component ID'] == 'HV_wrim_motor')
            ):
            str1 = 'Filtered_Excluded_Rule_Passed'     

            
        elif (str(row['split_rule']) == 'Excluded: {TS - Measure hard skirts} = No' 
             and (row['Measure hard skirts'] == 'No' )
             and (row['Strategy ID'] == 'M-SKIRT-INSP') 
             and (row['Component ID'] == 'skirt_hard')
            ):
            str1 = 'Skirt_Excluded_Rule_Passed'      
            
        elif (str(row['split_rule']) == 'Excluded: {TS - Magnet VVVF Powered} = "Y"' 
             and (row['Magnet VVVF Powered'] == 'Y' )
             and (row['Strategy ID'] == 'E-MAGNET-SVCE') 
             and (row['Component ID'] == 'magnet_conveyor')
            ):
            str1 = 'Magnet_VVF_Excluded_Rule_Passed'  
            
        elif (str(row['split_rule']) == 'Excluded: {TS - Magnet  has Travel Motor} = "N"' 
             and (row['Magnet  has Travel Motor'] == 'N' )
             and (row['Strategy ID'] == 'E-MAGNET-SVCE') 
             and (row['Component ID'] == 'magnet_conveyor')
            ):
            str1 = 'Magnet_No_Travel_Excluded_Rule_Passed'  

        elif (str(row['split_rule']) == 'Excluded: {TS - Primary scraper model} contains "Hosch".'
             and ('Hosch' in row['Primary scraper model'])
             and (row['Strategy ID'] == 'M-CLEANERS-INSP') 
             and (row['Component ID'] == 'scraper_primary')
            ):
            str1 = 'Hosch_Excluded_Rule_Passed'
        
        exclud.append(str1)
        
    exclud_logic_data['Exclusion_Criteria'] = exclud
    pass_data_for_excluded_tasks = exclud_logic_data[exclud_logic_data['Exclusion_Criteria']== 'Excluded_Rule_Keep']

    print('Task Excluded Logic Completed....')

    
    # Exlcusive logic ###############

    print('Running Task Exclusive Logic....')
    exclusiv_logic_data['Primary Scraper Tensioning Arrangement'] = exclusiv_logic_data['Primary Scraper Tensioning Arrangement'].str.lower().str.strip().astype(str)
    exclusiv_logic_data['Secondary Scraper Tensioning Arrangement'] = exclusiv_logic_data['Secondary Scraper Tensioning Arrangement'].str.lower().str.strip().astype(str)
    exclusiv_logic_data['Secondary Scraper Model'] = exclusiv_logic_data['Secondary Scraper Model'].str.lower().str.strip().astype(str)
    exclusiv_logic_data['Tertiary Scraper Tensioning Arrangement'] = exclusiv_logic_data['Tertiary Scraper Tensioning Arrangement'].str.lower().str.strip().astype(str)   
    exclusiv_logic_data['Tertiary Scraper '] = exclusiv_logic_data['Tertiary Scraper '].str.lower().str.strip().astype(str)



    exclusive = []

    for (i, row) in  exclusiv_logic_data.iterrows():
        
        shuttle_numbers = str(row['Shuttle pulleys']).split(',')
        pulley_number =  re.findall(r'^Pulley (\d+)', str(row['Description (from LMI)']))
        
        
        str1 = 'Not Exclusive'
        
        if ((str(row['split_rule']).find('Exclusive: Pulley numbers (from LMI list)') != -1 )
         and (any(number in pulley_number for number in shuttle_numbers) == True)):
            
            str1 = 'Pulley_Exclusive'   
            
        elif (str(row['split_rule']) == 'Exclusive: {TS - Bottom Cover BTT} = Yes' 
             and (row['Bottom Cover BTT'] == 'Yes')
             and (row['Strategy ID'] == 'M-BTT-OFF') 
             and (row['Component ID'] == 'belt_conveyor')
            ):
            str1 = 'Button_Cover_Exclusive'
        
        elif (str(row['split_rule']) == 'Exclusive: {TS - Gearbox permalubes} = "Y"'
             and (row['Gearbox Permalubes'] == 'Y' )
             and (row['Strategy ID'] == 'M-LUBE-ON') 
             and (row['Component ID'] == 'gearbox_conveyor')
            ):
            str1 = 'Gearbox_Exclusive'       
        
        elif (str(row['split_rule']) == 'Exclusive: {TS - Brush Wear Indicators} = Y'
             and (row['Brush Wear Indicators Present'] == 'Y' )
             and (row['Strategy ID'] == 'E-MOTR-INSP') 
             and (row['Component ID'] == 'HV_wrim_motor')
            ):
            str1 = 'Brush_Wear_Exclusive' 
            
        elif (str(row['split_rule']) == 'Exclusive: {TS - Drive Motor/s Included in General Inspection} = Y'
             and (row['Drive Motor/s Included in General Inspection'] == 'Y' )
             and (row['Strategy ID'] == 'E-CONV-SVCE' ) 
             and (row['Component ID'] == 'LV_motor_no-RTDs')
            ):
            str1 = 'Drive_Motor_Exclusive'
               
        elif (str(row['split_rule']) == 'Exclusive: {TS - Filters Replaced every 20Ws} = Y' 
             and (row['Filters Replaced every 20Ws'] == 'Y' )
             and (row['Strategy ID'] == 'E-MOTR-SVCE') 
             and (row['Component ID'] == 'HV_wrim_motor')
            ):
            str1 = 'Replace Filters_Exclusive' 
        
        elif (str(row['split_rule']) ==   'Exclusive: {TS - Site} = Whaleback' 
             and (row['Site'] == 'Whaleback' )
             and (row['Strategy ID'] == 'E-MOTR-SVCE') 
             and (row['Component ID'] == 'HV_wrim_motor')
            ):
            str1 = 'Site_Exclusive'
        
        elif (str(row['split_rule']) == 'Exclusive: {TS - Motor permalubes} = "Y"'
             and (row['Motor Permalubes'] == 'Y' )
             and (row['Strategy ID'] == 'M-LUBE-ON'| row['Strategy ID'] == 'E-MOTR-SVCE')
             and(row['Component ID']=='HV_wrim_motor'|row['Component ID'] == 'HV_scim_motor'|row['Component ID'] == 'LV_motor')
            ):
            str1 = 'MotorPerma_Exclusive'  
            
        elif (str(row['split_rule']) == 'Exclusive: {TS - Primary Scraper Tensioning Arrangement} contains "bolt".' 
             and ('bolt' in row['Primary Scraper Tensioning Arrangement'])
             and (row['Strategy ID'] == 'M-CLEANERS-INSP') 
             and (row['Component ID'] == 'scraper_primary')
            ):
            str1 = 'Primary_Bolt_Exclusive' 
        
        elif (str(row['split_rule']) == 'Exclusive: {TS - Primary Scraper Tensioning Arrangement} contains "spring".' 
             and ('spring' in row['Primary Scraper Tensioning Arrangement'])
             and (row['Strategy ID'] == 'M-CLEANERS-INSP') 
             and (row['Component ID'] == 'scraper_primary')
            ):
            str1 = 'Primary_Spring_Exclusive' 
            
        elif (str(row['split_rule']) == 'Exclusive: {TS - Primary Scraper Tensioning Arrangement} contains "J Bolt spring".' 
             and ('j bolt spring' in row['Primary Scraper Tensioning Arrangement'])
             and (row['Strategy ID'] == 'M-CLEANERS-INSP') 
             and (row['Component ID'] == 'scraper_primary')
            ):
            str1 = 'Primary_J_Bolt_Exclusive' 
            
        elif (str(row['split_rule']) == 'Exclusive: {TS - Primary Scraper Tensioning Arrangement} contains "HV-PCST Spring".' 
             and ('hv-pcst spring' in row['Primary Scraper Tensioning Arrangement'])
             and (row['Strategy ID'] == 'M-CLEANERS-INSP') 
             and (row['Component ID'] == 'scraper_primary')
            ):
            str1 = 'Primary_HV_PCST_Exclusive'
            
        elif (str(row['split_rule']) == 'Exclusive: {TS - Primary Scraper Tensioning Arrangement} contains "ESS XHD Spring".' 
             and ( 'ess xhd spring' in row['Primary Scraper Tensioning Arrangement'])
             and (row['Strategy ID'] == 'M-CLEANERS-INSP') 
             and (row['Component ID'] == 'scraper_primary')
            ):
            str1 = 'Primary_ESS_SPring_Exclusive'
            
        elif (str(row['split_rule']) == 'Exclusive: {TS - Secondary Scraper Tensioning Arrangement} contains "spring".' 
             and ('spring' in row['Secondary Scraper Tensioning Arrangement'])
             and (row['Strategy ID'] == 'M-CLEANERS-INSP') 
             and (row['Component ID'] == 'scraper_secondary')
            ):
            str1 = 'Secondary_Spring_Exclusive'
            
        elif (str(row['split_rule']) == 'Exclusive: {TS - Secondary Scraper Tensioning Arrangement} contains "SST".' 
             and ('sst' in row['Secondary Scraper Tensioning Arrangement'])
             and (row['Strategy ID'] == 'M-CLEANERS-INSP') 
             and (row['Component ID'] == 'scraper_secondary')
            ):
            str1 = 'SST_Spring_Exclusive'   
            
        elif (str(row['split_rule']) ==  'Exclusive: {TS - Secondary Scraper Tensioning Arrangement} contains "CST".' 
             and ('cst' in row['Secondary Scraper Tensioning Arrangement'] )
             and (row['Strategy ID'] == 'M-CLEANERS-INSP') 
             and (row['Component ID'] == 'scraper_secondary')
            ):
            str1 = 'CST_Spring_Exclusive'  
                    
        elif (str(row['split_rule']) == 'Exclusive: {TS - Secondary Scraper Tensioning Arrangement} contains "MST".' 
             and ('mst' in row['Secondary Scraper Tensioning Arrangement'])
             and (row['Strategy ID'] == 'M-CLEANERS-INSP') 
             and (row['Component ID'] == 'scraper_secondary')
            ):
            str1 = 'MST_Spring_Exclusive'  
            
        elif (str(row['split_rule']) == 'Exclusive: {TS - Secondary Scraper Model} contains "P Type".' 
             and ('p type' in row['Secondary Scraper Model'])
             and (row['Strategy ID'] == 'M-CLEANERS-INSP') 
             and (row['Component ID'] == 'scraper_secondary')
            ):
            str1 = 'Secondary_Scraper_P_Exclusive'
       
        elif (str(row['split_rule']) == 'Exclusive: {TS - Secondary Scraper Model} contains "R Type"' 
             and ('r type' in row['Secondary Scraper Model'])
             and (row['Strategy ID'] == 'M-CLEANERS-INSP') 
             and (row['Component ID'] == 'scraper_secondary')
            ):
            str1 = 'Secondary_Scraper_R_Exclusive'
            
        elif (str(row['split_rule']) == 'Exclusive: {TS - Tertiary Scraper Tensioning Arrangement} contains "SST".'
             and ('sst' in row['Tertiary Scraper Tensioning Arrangement'])
             and (row['Strategy ID'] == 'M-CLEANERS-INSP') 
             and (row['Component ID'] == 'scraper_tertiary')
            ):
            str1 = 'Tetiary_SST_Spring_Exclusive' 
     
        elif (str(row['split_rule']) ==  'Exclusive: {TS - Tertiary Scraper Tensioning Arrangement} contains "CST".' 
             and ('cst' in row['Tertiary Scraper Tensioning Arrangement'])
             and (row['Strategy ID'] == 'M-CLEANERS-INSP') 
             and (row['Component ID'] == 'scraper_secondary')
            ):
            str1 = 'Tetiarty_CST_Spring_Exclusive' 
            
        elif (str(row['split_rule']) ==   'Exclusive: {TS - Tertiary Scraper Tensioning Arrangement} contains "MST".' 
             and ('mst' in row['Tertiary Scraper Tensioning Arrangement'])
             and (row['Strategy ID'] == 'M-CLEANERS-INSP') 
             and (row['Component ID'] == 'scraper_tertiary')
            ):
            str1 = 'Tetiary_MST_Spring_Exclusive'   
                  
        elif (str(row['split_rule']) ==   'Exclusive: {TS - Tertiary Scraper Tensioning Arrangement} contains "spring".'
             and ('spring' in row['Tertiary Scraper Tensioning Arrangement'])
             and (row['Strategy ID'] == 'M-CLEANERS-INSP') 
             and (row['Component ID'] == 'scraper_tertiary')
            ):
            str1 = 'Tetiary_Spring_Exclusive' 

        elif (str(row['split_rule']) ==   'Exclusive: {TS - Tertiary Scraper Model} contains "P Type".'
             and ('p type' in row['Tertiary Scraper '])
             and (row['Strategy ID'] == 'M-CLEANERS-INSP') 
             and (row['Component ID'] == 'scraper_tertiary')
            ):
            str1 = 'Tetiary_Scraper_P_Exclusive'
            
        elif (str(row['split_rule']) ==    'Exclusive: {TS - Tertiary Scraper Model} contains "R Type"'
             and ('r type' in row['Tertiary Scraper '])
             and (row['Strategy ID'] == 'M-CLEANERS-INSP') 
             and (row['Component ID'] == 'scraper_tertiary')
            ):
            str1 = 'Tetiary_Scraper_R_Exclusive' 
            
        elif (str(row['split_rule']) ==    'Exclusive: {TS - Magnet VVVF Powered} = "Y"'
             and ('Y' in row['Magnet VVVF Powered'])
             and (row['Strategy ID'] == 'E-MAGNET-SVCE') 
             and (row['Component ID'] == 'magnet_conveyor')
            ):
            str1 = 'Magnet_VVF_Exclusive' 
         
        elif (str(row['split_rule']) ==    'Exclusive: {TS - Magnet Combo Pump / Fan} = "Y"'
             and ('Y' in row['Magnet Combo Pump / Fan'])
             and (row['Strategy ID'] == 'E-MAGNET-SVCE') 
             and (row['Component ID'] == 'magnet_conveyor')
            ):
            str1 = 'Magnet_Combo_Exclusive'
            
        elif (str(row['split_rule']) ==    'Exclusive: {TS - TMD Flag Drop Unit} = "Y"'
             and ('Y' in row['TMD Flag Drop Unit'])
             and (row['Strategy ID'] == 'E-TMD-SVCE') 
             and (row['Component ID'] == 'TMD')
            ):
            str1 = 'TMD_Flag_Exclusive'
            
        elif (str(row['split_rule']) ==    'Exclusive: {TS - TMD Spray Marker} = "Y"'
             and ('Y' in row['TMD Spray Marker'])
             and (row['Strategy ID'] == 'E-TMD-SVCE') 
             and (row['Component ID'] == 'TMD')
            ):
            str1 = 'TMD_Spray_Exclusive'
            
        elif (str(row['split_rule']) ==    'Exclusive: {TS - TMD Clip Detector} = "Y"'
             and ('Y' in row['TMD Clip Detector'])
             and (row['Strategy ID'] == 'E-TMD-SVCE') 
             and (row['Component ID'] == 'TMD')
            ):
            str1 = 'TMD_Clip_Exclusive'
            
        elif (str(row['split_rule']) ==    'Exclusive: {TS - TMD Flashing Beacon Light} = "Y"'
             and ('Y' in row['TMD Flashing Beacon Light'])
             and (row['Strategy ID'] == 'E-TMD-SVCE') 
             and (row['Component ID'] == 'TMD')
            ):
            str1 = 'TMD_Flashing_Beacon_Exclusive'
            
        elif (str(row['split_rule']) ==    'Exclusive: {TS - TMD High Pile Switch} = "Y"'
             and ('Y' in row['TMD High Pile Switch'])
             and (row['Strategy ID'] == 'E-TMD-SVCE') 
             and (row['Component ID'] == 'TMD')
            ):
            str1 = 'TMD_High_Pile_Exclusive'
        
        elif (str(row['split_rule']) == 'Exclusive: {TS - Secondary Scraper Tensioning Arrangement} contains "bolt" or "block".'
             and (('bolt' in row['Secondary Scraper Tensioning Arrangement'])|
                  ('block' in row['Secondary Scraper Tensioning Arrangement']))
             and (row['Strategy ID'] == 'M-CLEANERS-INSP') 
             and (row['Component ID'] == 'scraper_secondary')
            ):
            str1 = 'Secondary_Block_or_bolt_Exclusive'

        elif (str(row['split_rule']) == 'Exclusive: {TS - Tertiary Scraper Tensioning Arrangement} contains "bolt" OR "block"'
             and (('bolt' in row['Tertiary Scraper Tensioning Arrangement']) |
                  ('block' in row['Tertiary Scraper Tensioning Arrangement']))
             and (row['Strategy ID'] == 'M-CLEANERS-INSP') 
             and (row['Component ID'] == 'scraper_tertiary')
            ):
            str1 = 'Tetiary_Block_or_bolt_Exclusive'
            
        elif (str(row['split_rule']) == 'Exclusive: {TS - Tertiary Scraper Tensioning Arrangement} contains "bolt" OR "block"'
             and (('bolt' in row['Tertiary Scraper Tensioning Arrangement']) |
                  ('block' in row['Tertiary Scraper Tensioning Arrangement']))
             and (row['Strategy ID'] == 'M-CLEANERS-INSP') 
             and (row['Component ID'] == 'scraper_tertiary')
            ):
            str1 = 'Tetiary_Block_or_bolt_Exclusive'
        
        elif (str(row['split_rule']) == 'Exclusive: {TS - Take-up arrangement} contains "Festoon" '
             and ('Festoon' in row['Take-up arrangement'])
             and (row['Strategy ID'] == 'M-INSP-FPM') 
            ):
            str1 = 'Festoon_Exclusive'
            
        elif (str(row['split_rule']) == 'Exclusive: {TS - Take-up arrangement} contains "Trolley" '
             and ('Trolley' in row['Take-up arrangement'])
             and (row['Strategy ID'] == 'M-INSP-FPM') 
            ):
            str1 = 'Trolley_Exclusive'
            
        elif (str(row['split_rule']) == 'Exclusive: {TS - Take-up arrangement} contains "Winch" '
             and ('Winch' in row['Take-up arrangement'])
             and (row['Strategy ID'] == 'M-INSP-FPM') 
            ):
            str1 = 'Winch_Exclusive'
            
        elif (str(row['split_rule']) == 'Exclusive: {TS - Take-up arrangement} contains "Screw" '
             and ('Winch' in row['Take-up arrangement'])
             and (row['Strategy ID'] == 'M-INSP-FPM') 
            ):
            str1 = 'Screw_Exclusive'
        
        elif (str(row['split_rule']) == 'Exclusive: {TS - Take-up arrangement} contains "VVVF" '
             and ('Winch' in row['Take-up arrangement'])
             and (row['Strategy ID'] == 'M-INSP-FPM') 
            ):
            str1 = 'VVVF_Exclusive'
            
        elif (str(row['split_rule']) == 'Exclusive: {TS - Shuttle Rack or Chain} = "Rack" '
             and ('Rack' in str(row['Shuttle Rack or Chain']))
             and (row['Strategy ID'] == 'M-INSP-SHTL') 
            ):
            str1 = 'Rack_Exclusive'
            
        elif (str(row['split_rule']) ==  'Exclusive: {TS - Shuttle Rack or Chain} = "Chain" '
             and ('Chain' in str(row['Shuttle Rack or Chain']))
             and (row['Strategy ID'] == 'M-INSP-SHTL') 
            ):
            str1 = 'Chain_Exclusive'

        elif (str(row['split_rule']) ==  'Exclusive: {TS - PPC Tacho Wheel Proxy} = Y'
             and (row['PPC Tacho Wheel Proxy'] == 'Y')
             and (row['Strategy ID'] == 'E-WEIGHT-SVCE') 
            ):
            str1 = 'PPC_Tacho_Exclusive'
        
        elif (str(row['split_rule']) ==  'Exclusive: {TS - Pulley Speed Sensor} = Y'
             and (row['Pulley Speed Sensor'] == 'Y')
             and (row['Strategy ID'] == 'E-WEIGHT-SVCE') 
            ):
            str1 = 'Pulley_Speed_Exclusive'

        elif (str(row['split_rule']) ==   'Exclusive: {TS - TES Project} = Port' 
             and  ('Port' in row['TES Project'])
             and (row['Strategy ID'] == 'E-WEIGHT-SVCE') 
             and (row['Component ID'] == 'weightometer')
            ):
            str1 = 'Port_Exclusive'

        elif (str(row['split_rule']) == 'Exclusive: {TS - Primary scraper model} contains "Hosch".'
             and ('Hosch' in row['Primary scraper model'])
             and (row['Strategy ID'] == 'M-CLEANERS-INSP') 
             and (row['Component ID'] == 'scraper_primary')
            ):
            str1 = 'Hosch_Exclusive'
            
        exclusive.append(str1)
               
    exclusiv_logic_data['Exclusion_Criteria'] = exclusive

    pass_exclusiv_logic_data = exclusiv_logic_data[exclusiv_logic_data['Exclusion_Criteria'] != 'Not Exclusive']   
    
    print('Completed Task Exclusive Logic....')



    # #### `Reference Images:`

    imgR = []

    for (i, row) in  pass_exclusiv_logic_data.iterrows():
        
        str1 = str(row['Reference images']).replace('{TS - Secondary Scraper Tensioning Arrangement Image}.', 
                         (str(row['Secondary Scraper Tensioning Arrangement Image']))).replace('{TS - Tertiary Scraper Tensioning Arrangement Image}.', 
                         str(row['Tertiary Scraper Tensioning Arrangement Image']))
            
        imgR.append(str1) 
        
        
    pass_exclusiv_logic_data['Reference images'] = imgR


    # ### `Stitch the Data back together:`


    cols = ['Exclusion_Criteria']

    stitch_data_1 =  pass_exclusiv_logic_data.drop(cols, axis =1)
    stitch_data_2 =  pass_data_for_excluded_tasks.drop(cols, axis =1)

    wsws_3a = pass_data.append( [stitch_data_1,stitch_data_2], ignore_index= True)
    print('Data Stitched back together....')

    
    component_library_2 = pd.read_excel(file2,sheet_name= 'Strategy & PMI info', skiprows= range(0,1))
    cols =['Strategy ID', 'In Scope', 'Description' , 'PRT Title', 'System condition']
    component_library_2[cols] = component_library_2[cols].fillna(method= 'ffill')
    component_library_2.rename(columns={'Strategy':'Strategy ID', 'Description': 'Description (from CL)'}, inplace=True)
    wsws_3b = pd.merge(wsws_3a, component_library_2, on= ['Strategy ID'], how= 'inner')
    wsws_3c= wsws_3b[wsws_3b['In Scope']=='Yes']
    wsws_3c['Freq_Week'] = wsws_3c['Frequency'] + ':' +  wsws_3c['Frequency_In_Weeks']
    wsws_3c['Stra-Conv'] =   wsws_3c['Strategy ID'] + '-' + wsws_3c['Conveyor'] 
    wsws_3= wsws_3c


    print('Tasks Stacking in progress...')

    wsws_3['Frequency_In_Weeks'] = wsws_3['Frequency_In_Weeks'].str.replace('nan','0W' )
    stacker = wsws_3.groupby(wsws_3['Stra-Conv'], as_index=True )['Freq_Week'].unique().to_frame()
    stacker['Stra-Conv'] = stacker.index
    stacker.set_index(['Stra-Conv'], drop =True)
    stacker['FW'] = stacker['Freq_Week']
    stacker_df1 = stacker[['Stra-Conv', 'FW']].reset_index(drop =True)
    wsws_4 = pd.merge(stacker_df1, wsws_3, on= ['Stra-Conv'], how= 'inner')
    wsws_5 = wsws_4.explode('FW')
    wsws_5[['New_Level','New_Freq']] = wsws_5['FW'].str.split(":",expand=True,)
    wsws_5['Original'] = wsws_5.apply(lambda x: '0' if 
      x['Frequency_In_Weeks'] == str(np.nan)
      else x['Frequency_In_Weeks'], axis=1).str.extract('(\d+)').fillna(0).astype(int)

    wsws_5['Stacked'] = wsws_5.apply(lambda x: '0' if 
      x['New_Freq'] == str(np.nan)
      else x['New_Freq'], axis=1).str.extract('(\d+)').fillna(0).astype(int)

    wsws_5a = wsws_5[wsws_5['Original'] <= wsws_5['Stacked']]
    wsws_6 =  wsws_5a.drop(columns=['Original', 'Stacked', 'FW', 'Frequency', 'Frequency_In_Weeks'])
    wsws_9=  wsws_6.drop_duplicates().reset_index(drop = True)


    print('Stacking Completed.....')
    

    print('Logic for WSWS running....')

    ## Logic For WSWS
    ### `Create DMS, DIR and PRT : `

    # #### `DMS : `

    # In[24]:


    dms= []

    for (i, row) in  wsws_9.iterrows():
        
        if  str(row['Strategy ID']).find('M-OIL-SMPL') != -1:
            
            str1 = str(row['Strategy ID']) + '-'+ str(row['Site']) +     '-'+  str(row['Oil sampling document circuit'])+'-' +   str(row['New_Freq'])
        
    
        elif  str(row['Strategy ID']).find('E-STAT-SIREN') != -1:
        
            str1 = str(row['Strategy ID']) + '-'+ str(row['Site']) +  '-'+  str(row['Siren Stat MP circuits'])+'-' +   str(row['New_Freq'])
        
        elif  str(row['Strategy ID']).find('M-VA-SEV2') != -1:

            str1 = str(row['Strategy ID']) + '-'+ str(row['Site']) + '-'+  str(row['VA Severity 2 Document Circuit'])+ '-' +   str(row['New_Freq'])

        elif  str(row['Strategy ID']).find('E-WEIGHT-SVCE') != -1:
        
            str1 = str(row['Strategy ID']) + '-'+ str(row['Site']) + '-'+ str(row['Conveyor'])+ '-'+  str(row['Description (from LMI)']) + '-' + str(row['New_Freq'])

        else: 
            str1 = str(row['Strategy ID']) + '-'+ str(row['Site']) +     '-'+  str(row['Conveyor'])+'-' +   str(row['New_Freq'])
            
        dms.append(str1)
        
    wsws_9['DMS (To-Be)'] = dms


    # #### `PRT ShortTEXT : `

    # In[25]:


    prt = []

    for (i, row) in  wsws_9.iterrows():
        
        if  str(row['Strategy ID']).find('M-OIL-SMPL') != -1:
            
            str1 = str(row['New_Freq']) + ' ' + str(row['PRT Title'])+ ', ' +  str(row['Oil sampling document circuit'])

        elif  str(row['Strategy ID']).find('E-STAT-SIREN') != -1:
        
            str1 = str(row['New_Freq']) + ' ' + str(row['PRT Title'])+ ', ' +  str(row['Siren Stat MP circuits'])
        
        elif  str(row['Strategy ID']).find('M-VA-SEV2') != -1:

            str1 = str(row['New_Freq']) + ' ' + str(row['PRT Title'])+ ', ' + str(row['VA Severity 2 Document Circuit'])
  
        elif  str(row['Strategy ID']).find('E-WEIGHT-SVCE') != -1:
        
            str1 = str(row['New_Freq']) + ' ' + str(row['PRT Title'])+ ', ' + str(row['Description (from LMI)'])
            
        else: 
            str1 = str(row['New_Freq']) + ' ' + str(row['PRT Title'])+ ', ' + str(row['Conveyor'])
             
        prt.append(str1)
        
    wsws_9['PRT Short Text (To-Be)'] = prt


    # #### `DIR and PRT Task Sequence:`

    # In[26]:


    wsws_9['DIR to be'] = 'DMS-Group-' + wsws_9.groupby(['DMS (To-Be)']).ngroup().astype(str) 
    wsws_9['DIR to be'].nunique()
    wsws_9['PRT Task # (Sequence) to be'] = wsws_9['Task order']


    ### `PMI Documents Logics`

    # #### `Document Picker:`

    # In[27]:


    doc = []
    desc = []
    fname = []
    excluder = []

    for (i, row) in  wsws_9.iterrows():

        if row['Site']  == 'Whaleback':
            str1 = str(row['MWB Reference Documentation (To-Be)\n - Doc#']).replace('{TS - Primary Replace WIN No.}',
               str(row['Primary Replace WIN No.'])).replace('{TS - Secondary Replace WIN No.}',
               str(row['Secondary Replace WIN No.'])).replace('{TS - Tertiary Replace WIN No.}',
               str(row['Tertiary Replace WIN No.'])).replace('{TS - Primary Refurb WIN No.}',
               str(row['Primary Refurb WIN No.'])).replace('{TS - Secondary Refurb WIN No.}',
               str(row['Secondary Refurb WIN No.'])).replace('{TS - Tertiary Refurb WIN No.}',
               str(row['Tertiary Refurb WIN No.'])).replace('{TS - Shuttle Encoder WIN No.}',
               str(row['Shuttle Encoder WIN No.']))
            
            
            
            str2 = str(row['MWB Reference Documentation (To-Be)\n - Description']).replace('{TS - Primary Replace WIN Title}',
               str(row['Primary Replace WIN Title'])).replace('{TS - Secondary Replace WIN Title}',
               str(row['Secondary Replace WIN Title'])).replace('{TS - Tertiary Replace WIN Title}',
               str(row['Tertiary Replace WIN Title'])).replace('{TS - Primary Refurb WIN Title}',
               str(row['Primary Refurb WIN Title'])).replace('{TS - Secondary Refurb WIN Title}',
               str(row['Secondary Refurb WIN Title'])).replace('{TS - Tertiary Refurb WIN Title}',
               str(row['Tertiary Refurb WIN Title'])).replace('{TS - Shuttle Encoder WIN Title}',
               str(row['Shuttle Encoder WIN Title']))
            
            
            str3 = str(row['MWB Reference Documentation (To-Be)\n-File Name'])
            str4 = str(row['EXC. DNMWB'])
        
        if row['Site']  == 'Port':
            str1 = str(row['PORT Reference Documentation (To-Be)\n - Doc#'])
            str2 = str(row['PORT Reference Documentation (To-Be)\n - Description'])
            str3 = str(row['PORT Reference Documentation (To-Be)\n-File Name'])
            str4 = str(row['EXC. PORT'])
            
        if row['Site']  == 'Finucane':
            str1 = str(row['FINUCANE ISLAND Reference Documentation (To-Be)\n - Doc#'])
            str2 = str(row['FINUCANE ISLAND Reference Documentation (To-Be)\n - Description'])
            str3 = str(row['FINUCANE ISLAND Reference Documentation (To-Be)\n-File Name'])
            str4 = str(row['EXC. DNFI'])
         
        if row['Site']  == 'Nelson Point':
            str1 = str(row['NELSON POINT Reference Documentation (To-Be)\n - Doc#'])
            str2 = str(row['NELSON POINT Reference Documentation (To-Be)\n - Description'])
            str3 = str(row['NELSON POINT Reference Documentation (To-Be)\n-File Name'])
            str4 = str(row['EXC. DNNP'])
            
        doc.append(str1)
        desc.append(str2)
        fname.append(str3)
        excluder.append(str4)
            
    wsws_9['Doc'] = doc
    wsws_9['FDesc'] = desc
    wsws_9['Filename'] = fname
    wsws_9['ExcludeDocs'] = excluder


    # #### `Image Name Replacer:`

    # In[28]:


    img = []

    for (i, row) in  wsws_9.iterrows():
        str1 = str(row['Filename']).replace('{TS - Pulley diagram}', str(row['Pulley diagram'])).replace('{TS - Air Gap Adjustment}', str(row['Air Gap Adjustment'])).replace('{TS - Fluid Coupling Schematic}', str(row['Fluid Coupling Schematic']))
        
            
        img.append(str1)
        

    wsws_9['Filename'] = img
    wsws_9['Filename'].unique()


    # In[29]:


    wsws_9['Reference images'].unique()


    # #### `Document to be Replacer:`

    # In[30]:


    docR = []
    fileR = []
    for (i, row) in wsws_9.iterrows():
        
        
        str1 = str(row['Doc']).replace('{TS - Primary Replace WIN No.}',
               str(row['Primary Replace WIN No.'])).replace('{TS - Primary Replace WIN Title}',
               str(row['Primary Replace WIN Title'])).replace('{TS - Secondary Replace WIN No.}',
               str(row['Secondary Replace WIN No.'])).replace('{TS - Secondary Replace WIN Title}',
               str(row['Secondary Replace WIN Title'])).replace('{TS - Tertiary Replace WIN No.}',
               str(row['Tertiary Replace WIN No.'])).replace('{TS - Tertiary Replace WIN Title}',
               str(row['Tertiary Replace WIN Title'])).replace('{TS - Primary Refurb WIN No.}',
               str(row['Primary Refurb WIN No.'])).replace('{TS - Primary Refurb WIN Title}',
               str(row['Primary Refurb WIN Title'])).replace('{TS - Secondary Refurb WIN No.}',
               str(row['Secondary Refurb WIN No.'])).replace('{TS - Secondary Refurb WIN Title}',
               str(row['Secondary Refurb WIN Title'])).replace('{TS - Tertiary Refurb WIN No.}',
               str(row['Tertiary Refurb WIN No.'])).replace('{TS - Tertiary Refurb WIN Title}',
               str(row['Tertiary Refurb WIN Title']))
        
        str2 = str(row['FDesc']).replace('{TS - Primary Replace WIN No.}',
               str(row['Primary Replace WIN No.'])).replace('{TS - Primary Replace WIN Title}',
               str(row['Primary Replace WIN Title'])).replace('{TS - Secondary Replace WIN No.}',
               str(row['Secondary Replace WIN No.'])).replace('{TS - Secondary Replace WIN Title}',
               str(row['Secondary Replace WIN Title'])).replace('{TS - Tertiary Replace WIN No.}',
               str(row['Tertiary Replace WIN No.'])).replace('{TS - Tertiary Replace WIN Title}',
               str(row['Tertiary Replace WIN Title'])).replace('{TS - Primary Refurb WIN No.}',
               str(row['Primary Refurb WIN No.'])).replace('{TS - Primary Refurb WIN Title}',
               str(row['Primary Refurb WIN Title'])).replace('{TS - Secondary Refurb WIN No.}',
               str(row['Secondary Refurb WIN No.'])).replace('{TS - Secondary Refurb WIN Title}',
               str(row['Secondary Refurb WIN Title'])).replace('{TS - Tertiary Refurb WIN No.}',
               str(row['Tertiary Refurb WIN No.'])).replace('{TS - Tertiary Refurb WIN Title}',
               str(row['Tertiary Refurb WIN Title']))
        
        docR.append(str1)
        fileR.append(str2)
        
    wsws_9['Doc'] = docR
    wsws_9['FDesc'] = fileR


    # #### `Equipment and Tooling quantities column Replacer:`

    # In[31]:


    wsws_9['Number of Drives_int'] =  (wsws_9['Number of Drives'].astype(float)) * 2

    non_perma = []

    for (i, row) in  wsws_9.iterrows():
        
        str1 = str(row['Equipment, Special Tooling & Consumables Required\n - Qty']).replace('{TS - Non-shuttle Perma qty}',     str(row['Non-shuttle perma qty'])).replace('{TS - Shuttle Perma qty}',     str(row['Shuttle perma qty'])).replace('{TS - Breather qty}',     str(row['Breather qty'])).replace('{TS - Smpl kit qty}',     str(row['Smple kit qty'])).replace('{TS - Number of Drives}*2',     str(row['Number of Drives_int']))
        
            
        non_perma.append(str1)
        

    wsws_9['Equipment, Special Tooling & Consumables Required\n - Qty'] = non_perma
    #wsws11['Equipment, Special Tooling & Consumables Required\n - Qty'] = wsws11['Equipment, Special Tooling & Consumables Required\n - Qty'].astype(str)
    wsws_9['Equipment, Special Tooling & Consumables Required\n - Qty'] = wsws_9['Equipment, Special Tooling & Consumables Required\n - Qty'].str.split('.').str[0].str.replace('nan', '0')

    


    #### `Equipment, Special Tooling & Consumables Required - Resource Description column Replacer:`

    resource_replacer = []

    for (i, row) in wsws_9.iterrows():

        str1 = str(row['Equipment, Special Tooling & Consumables Required \n - Resource Description']).replace('{TS - Accumulator Pre-charge}', str(row['Accumulator Pre-charge']))


        resource_replacer.append(str1)
    
    
    wsws_9['Equipment, Special Tooling & Consumables Required \n - Resource Description'] = resource_replacer
    

    # ### `Task Supression:`

    # In[32]:

    print('Supression In Progress...!!!')

    wsws_9['Supresses'] = wsws_9['Supresses'].astype(str).str.strip().str.replace(' ', '')
    wsws_9['Task order'] = wsws_9['Task order'].astype(str).str.replace(' ', '')

    wsws_9['SupressTag'] = wsws_9['DMS (To-Be)'] +'-'+ wsws_9['Description (from LMI)'] + '-'
    wsws_9['SupressMe'] = wsws_9['DMS (To-Be)'] +'-'+ wsws_9['Description (from LMI)'] +'-'+ wsws_9['Task order'].astype(str)


    wsws10 = wsws_9.drop_duplicates(subset = ['DIR to be',
         'PRT Task # (Sequence) to be', 'Description (from LMI)',
          'Task description'], keep = 'first').reset_index(drop=True)

    stx = []
    for (i, row) in wsws10.iterrows():
        
        str1 = row['Supresses'].strip().split(',')
        
        str2 = list(map(lambda x: row['SupressTag'] + x, str1))
        
        stx.append(str2)
        
    wsws10['Supressor'] = stx

    suppres_list = [_ for sublist in wsws10['Supressor'].to_list() for _ in sublist]

    rm = []
    for _,row in wsws10.iterrows():
        str1 = 'Keep'
        if row['SupressMe'] in suppres_list:
            str1 = 'Remove'

        rm.append(str1)
        
    wsws10['Condition'] = rm
    wsws11 = wsws10[wsws10['Condition'] == 'Keep']
    print('Supression Completed...!!!')

    
    #### `Create New DIR:`
    wsws11['DIR to be'] = 'DMS-Group-' + wsws11.groupby(['DMS (To-Be)']).ngroup().astype(str) 
    
    ##  Creating Seprate PRT and PMI Sheets:
    pmi_excluded = wsws_9[['DMS (To-Be)','Site', 'DIR to be', 'PRT Short Text (To-Be)', 'Strategy ID', 'In Scope', 'Description (from CL)', 'PRT Title',
       'System condition', 'PMI Header Image Description', 'EXC. ID',
       'PMI Header Image Filename', 'EXC. IF', 'Scope', 'EXC. Sco',
       'Additional Safety Statements', 'EXC. ASS',
       'Additional Controls - Control Type', 'EXC. CT',
       'Additional Controls\n - Reason for Control Requirements', 'EXC. RCR',
       'Additional PPE ', 'EXC. PPE',
       'Equipment, Special Tooling & Consumables Required\n - Resource Type',
       'EXC. RT',
       'Equipment, Special Tooling & Consumables Required \n - Resource Description',
       'EXC. RD', 'Equipment, Special Tooling & Consumables Required\n - Qty',
       'EXC. Qty', 'Resource Special Skills (To-Be)', 'EXC. RSS',
       'NELSON POINT Reference Documentation (To-Be)\n - Doc#', 'EXC. DNNP',
       'NELSON POINT Reference Documentation (To-Be)\n - Description',
       'EXC. DDNP', 'NELSON POINT Reference Documentation (To-Be)\n-File Name',
       'EXC. DFNNP',
       'FINUCANE ISLAND Reference Documentation (To-Be)\n - Doc#', 'EXC. DNFI',
       'FINUCANE ISLAND Reference Documentation (To-Be)\n - Description',
       'EXC. DDFI',
       'FINUCANE ISLAND Reference Documentation (To-Be)\n-File Name',
       'EXC. DFNFI', 'MWB Reference Documentation (To-Be)\n - Doc#',
       'EXC. DNMWB', 'MWB Reference Documentation (To-Be)\n - Description',
       'EXC. DDMWB', 'MWB Reference Documentation (To-Be)\n-File Name',
       'EXC. DFNMWB', 'Handover Activities (To-Be)', 'EXC. HOA',
       'Housekeeping Activities (To-Be)', 'EXC. HKA', 'Doc', 'FDesc', 'Filename', 'New_Level' , 'Default Sequencing',
        'ExcludeDocs'        
                      ]]
#pmi_sheet_1 = pmi_excluded.drop_duplicates(subset = pmi_excluded.columns.difference(['Name'])).reset_index(drop=True)

    pmi_sheet_1 = pmi_excluded.drop_duplicates().reset_index(drop=True)

    ## ` PRT Sheet:`
    dict_prt = ['Unique #', 'Notes', 'MaintenancePlan', 'MaintItem', 'TL#2',
       'Task List Description', 'DMS (To-Be)', 'DIR to be', 'In To-Be',
       'PRT Short Text (To-Be)', 'LMI (To-Be)', 'LMI QA', 'Tag Number (As-Is)',
       'Asset Strategy Template L1', 'Asset Strategy Template L2',
       'Asset Strategy Template L3', 'Task Interval (To-Be)',
       'Task Interval Hrs (To-Be)', 'PRT System Condition',
       'PRT System Condition Text', 'PRT Task # (Sequence) to be',
       'Task Description to be', 'QA', 'Acceptable Limit to be', 'QA2',
       'Contains Materials to be', 'Material No (To-Be)',
       'Material Description (To-Be)', 'Qty (To-Be)', 'Unit (To-Be)',
       'OEM# (To-Be)', 'GRC ID (To-Be)', 'GRC Description (To-Be)',
       'CCE (To-Be)', 'CCV (To-Be)', 'Task Hazard Symbol (To-Be)',
       'Measurement Box (TRUE/FALSE) (To-Be)', 'Corrective Action (To-Be)',
       'Corrective Task List Description (To-Be)',
       'Resource Work Centre (RWC) (To-Be)',
       'Resource Work Centre Description (To-Be)',
       'OnePM Primary Labour (To-Be)', 'PMI Execution Condition',
       'PMI Execution Condition Statement',
       'Activity Image Filname/Path (To-Be)',
       'OnePM Package Activity Type (To-Be)', 'OnePM Package Type Group' ,'Default Sequencing','Function',
       'Functional Failure', 'Failure Mode', 'Failure Detectability',
       'Strategy Selection', 'Activity Selection']

    prt_sheet  = pd.DataFrame(columns= dict_prt)
    prt_sheet['DMS (To-Be)'] = wsws11['DMS (To-Be)']
    prt_sheet['DIR to be'] = wsws11['DIR to be']
    prt_sheet['In To-Be'] = 'IN'
    prt_sheet['PRT Short Text (To-Be)'] = wsws11['PRT Short Text (To-Be)']
    prt_sheet['LMI (To-Be)'] = wsws11['Description (from LMI)'].astype(str)
    prt_sheet['PRT Task # (Sequence) to be'] = wsws11['Task order']
    prt_sheet['Task Description to be'] = wsws11['Task description']
    prt_sheet['Task Description to be'] = wsws11['Task description']
    prt_sheet['Acceptable Limit to be'] = wsws11['Acceptable limit']
    prt_sheet['Contains Materials to be'] = wsws11['Contains Materials (TRUE/FALSE)']
    prt_sheet['GRC ID (To-Be)'] = wsws11['GRC ID']
    prt_sheet['GRC Description (To-Be)'] = wsws11['GRC Description']
    prt_sheet['Measurement Box (TRUE/FALSE) (To-Be)']= wsws11['Measurement Box (TRUE/FALSE)']
    prt_sheet['Corrective Action (To-Be)'] = wsws11['Corrective action']
    prt_sheet['Resource Work Centre (RWC) (To-Be)']= wsws11['Resource Work Centre (RWC)']
    prt_sheet['PMI Execution Condition'] = wsws11['PMI Execution Condition']
    prt_sheet['Activity Image Filname/Path (To-Be)'] = wsws11['Reference images']
    prt_sheet['Default Sequencing'] = wsws11['Default Sequencing']
    prt_sheet['NameSort'] = wsws11['Name'].astype(str)
    prt_sheet['Function'] = wsws11['Function']
    prt_sheet['Functional Failure'] = wsws11['Functional Failure']
    prt_sheet['Failure Mode'] = wsws11['Failure Mode']
    prt_sheet['Failure Detectability'] = wsws11['Failure Detectability']
    prt_sheet['Strategy Selection'] = wsws11['Strategy Selection']
    prt_sheet['Activity Selection'] = wsws11['Activity Selection']
    prt_sheet_2 = prt_sheet.drop_duplicates().reset_index(drop= True)
    prt_sheet_3 = prt_sheet_2.sort_values(by = ['Default Sequencing'], ascending= True)

    # #### `Sorter Logics :`

    # In[35]:


    sort = []

    for (i, row) in  prt_sheet_3.iterrows():
        
        str1 = row['Default Sequencing'][:3]
        str2 = re.sub(".*\D", '', str1)
        str3 = pd.to_numeric(str2, errors='coerce')
        sort.append(str3)
        
        
    prt_sheet_3['SorterX'] = sort
    prt_sheet_4 = prt_sheet_3.sort_values(by = ['DIR to be', 'SorterX'])


    # #### #### `Unique Row # :`


    row_id = 1
    u_row = []
    for (i, row) in prt_sheet_4.iterrows():
        str1 =  str(row_id)
        row_id = row_id + 1
        u_row.append(str1)

    prt_sheet_4['Unique #'] = u_row


    cols = ['Default Sequencing', 'SorterX', 'NameSort']
    final_prt = prt_sheet_4.drop(cols, axis =1)
    print('Final PRT Sheet Created...')

    # #### ` PMI Sheet:`

    # In[38]:

    print('Creating Final PMI Sheets...')
    dict_pmi = ['DMS# (To-Be)', 'DIR# (To-Be)', 'Version (To-Be)',
           'PRT Work Description Short Text (Document Title) (To-Be)',
           'PMI Header Image Description (To-Be)',
           'PMI Header Image Filename (To-Be)', 'Scope',
           'Additional Safety Statements (To-Be)',
           'Additional Controls (To-Be)\n - Control Type',
           'Additional Controls (To-Be)\n - Reason for Control Requirements',
           'Additional PPE (To-Be)',
           'Equipment, Special Tooling & Consumables Required (To-Be)\n - Resource Type',
           'Equipment, Special Tooling & Consumables Required (To-Be)\n - Resource Description',
           'Equipment, Special Tooling & Consumables Required (To-Be)\n - Qty',
           'Resource Special Skills (To-Be)',
           'Reference Documentation (To-Be)\n - Doc#',
           'Reference Documentation (To-Be)\n - Description',
           'Reference Documentation (To-Be)\n-File Name',
           'Handover Activities (To-Be)', 'Housekeeping Activities (To-Be)']
    pmi_sheet_2 = pd.DataFrame(columns= dict_pmi)


    # In[39]:


    pmi_sheet_2['DMS# (To-Be)'] = pmi_sheet_1['DMS (To-Be)']
    pmi_sheet_2['DIR# (To-Be)'] = pmi_sheet_1['DIR to be']
    pmi_sheet_2['PRT Work Description Short Text (Document Title) (To-Be)'] = pmi_sheet_1['PRT Short Text (To-Be)']
    pmi_sheet_2['PMI Header Image Description (To-Be)'] = pmi_sheet_1['PMI Header Image Description']
    pmi_sheet_2['PMI Header Image Filename (To-Be)'] = pmi_sheet_1['PMI Header Image Filename']
    pmi_sheet_2['Scope'] = pmi_sheet_1['Scope'] 
    pmi_sheet_2['Additional Safety Statements (To-Be)'] = pmi_sheet_1['Additional Safety Statements'] 
    pmi_sheet_2['Additional Controls (To-Be)\n - Control Type'] = pmi_sheet_1['Additional Controls - Control Type'] 
    pmi_sheet_2['Additional Controls (To-Be)\n - Reason for Control Requirements'] = pmi_sheet_1['Additional Controls\n - Reason for Control Requirements'] 
    pmi_sheet_2['Additional PPE (To-Be)'] = pmi_sheet_1['Additional PPE '] 
    pmi_sheet_2['Equipment, Special Tooling & Consumables Required (To-Be)\n - Resource Type'] =pmi_sheet_1['Equipment, Special Tooling & Consumables Required\n - Resource Type'] 
    pmi_sheet_2['Equipment, Special Tooling & Consumables Required (To-Be)\n - Resource Description'] = pmi_sheet_1['Equipment, Special Tooling & Consumables Required \n - Resource Description']
    pmi_sheet_2['Equipment, Special Tooling & Consumables Required (To-Be)\n - Qty'] = pmi_sheet_1['Equipment, Special Tooling & Consumables Required\n - Qty'] 
    pmi_sheet_2['Resource Special Skills (To-Be)'] = pmi_sheet_1['Resource Special Skills (To-Be)'] 
    pmi_sheet_2['Reference Documentation (To-Be)\n - Doc#'] = pmi_sheet_1['Doc'].str.replace('nan', "N/A")
    pmi_sheet_2['Reference Documentation (To-Be)\n - Description'] = pmi_sheet_1['FDesc'] 
    pmi_sheet_2['Reference Documentation (To-Be)\n-File Name'] = pmi_sheet_1['Filename'] 
    pmi_sheet_2['Handover Activities (To-Be)'] = pmi_sheet_1['Handover Activities (To-Be)'] 
    pmi_sheet_2['Housekeeping Activities (To-Be)'] = pmi_sheet_1['Housekeeping Activities (To-Be)'] 
    pmi_sheet_2['EXC. RT'] = pmi_sheet_1['EXC. RT']
    pmi_sheet_2['EXC. RD'] = pmi_sheet_1['EXC. RD']
    pmi_sheet_2['EXC. Qty'] = pmi_sheet_1['EXC. Qty']
    pmi_sheet_2['Freq_exclusion'] = pmi_sheet_1['New_Level']
    pmi_sheet_2['Doc_exclusion'] = pmi_sheet_1['ExcludeDocs']
    pmi_sheet_2['Default Sequencing'] = pmi_sheet_1['Default Sequencing'] 


    # #### `Logic for PMI Exclusion Levels:`

    # In[40]:


    rt  = []
    rd  = []
    qty = []
    for (i, row) in pmi_sheet_2.iterrows():
        
        list1 = str(row['Freq_exclusion']).split(',')
        list2 = str(row['EXC. RT']).split(', ')
        list3 = str(row['EXC. RD']).split(', ')
        list4 = str(row['EXC. Qty']).split(', ')
   
        
        str1 =  str(row['Equipment, Special Tooling & Consumables Required (To-Be)\n - Resource Type'])
        str2 =  str(row['Equipment, Special Tooling & Consumables Required (To-Be)\n - Resource Description'])
        str3 =  str(row['Equipment, Special Tooling & Consumables Required (To-Be)\n - Qty'])
     
        if list(set(list1).intersection(set(list2))) != []:
            str1 =  ''
            
        if list(set(list1).intersection(set(list3))) != []:
            str2 =  '' 
            
        if list(set(list1).intersection(set(list4))) != []:
            str3 =  ''   
            
        rt.append(str1)
        rd.append(str2)
        qty.append(str3)

    pmi_sheet_2['Equipment, Special Tooling & Consumables Required (To-Be)\n - Resource Type'] = rt
    pmi_sheet_2['Equipment, Special Tooling & Consumables Required (To-Be)\n - Resource Description'] = rd
    pmi_sheet_2['Equipment, Special Tooling & Consumables Required (To-Be)\n - Qty'] = qty


    # ##### `Logic for PMI Document Levels:`

    # In[41]:

    #### `Remove PMI Reference Document  Frequency Levels :`
    pmi_sheet_2['Freq_exclusion'] = pmi_sheet_2['Freq_exclusion'].astype(str).str.strip().str.replace(' ', '')
    pmi_sheet_2['Doc_exclusion'] = pmi_sheet_2['Doc_exclusion'].astype(str).str.strip().str.replace(' ', '')

    pmi_sheet_2['Freq_exclusion'] = pmi_sheet_2['Freq_exclusion'].astype(str).str.strip().str.replace(' ', '')
    pmi_sheet_2['Doc_exclusion'] = pmi_sheet_2['Doc_exclusion'].astype(str).str.strip().str.replace(' ', '')

    doc  = []
    desc  = []
    fname = []

    for (i, row) in pmi_sheet_2.iterrows():

        str1 =  str(row['Reference Documentation (To-Be)\n - Doc#'])
        str2 =  str(row['Reference Documentation (To-Be)\n - Description'])
        str3 =  str(row['Reference Documentation (To-Be)\n-File Name'])


        if row['Doc_exclusion'].find(row['Freq_exclusion']) != -1:

            str1 =  ''
            str2 =  '' 
            str3 =  ''   

        doc.append(str1)
        desc.append(str2)
        fname.append(str3)

    pmi_sheet_2['Reference Documentation (To-Be)\n - Doc#'] = doc
    pmi_sheet_2['Reference Documentation (To-Be)\n - Description'] = desc
    pmi_sheet_2['Reference Documentation (To-Be)\n-File Name'] = fname


    #### `PPE Duplication Fix :`
    pmi_sheet_2.loc[(pmi_sheet_2['DIR# (To-Be)'].duplicated() & pmi_sheet_2['Additional PPE (To-Be)'].duplicated()), ['Additional PPE (To-Be)']] = ''
    pmi_sheet_2 = pmi_sheet_2.copy()
    # #### `Logic for Equipment QTY Counter :`

    # In[42]:


    rt  = []
    rd  = []
    qty = []
    for (i, row) in pmi_sheet_2.iterrows():
        
       
        
        str1 =  str(row['Equipment, Special Tooling & Consumables Required (To-Be)\n - Resource Type'])
        str2 =  str(row['Equipment, Special Tooling & Consumables Required (To-Be)\n - Resource Description'])
        str3 =  str(row['Equipment, Special Tooling & Consumables Required (To-Be)\n - Qty'])
        
        
     
        if row['Equipment, Special Tooling & Consumables Required (To-Be)\n - Qty']== '0':
            
            str1 =  ''
            str2 =  '' 
            str3 =  ''   
            
        rt.append(str1)
        rd.append(str2)
        qty.append(str3)

    pmi_sheet_2['Equipment, Special Tooling & Consumables Required (To-Be)\n - Resource Type'] = rt
    pmi_sheet_2['Equipment, Special Tooling & Consumables Required (To-Be)\n - Resource Description'] = rd
    pmi_sheet_2['Equipment, Special Tooling & Consumables Required (To-Be)\n - Qty'] = qty
    cols = ['EXC. RT', 'EXC. RD', 'EXC. Qty','Freq_exclusion', 'Default Sequencing', 'Doc_exclusion']
    final_pmi = pmi_sheet_2.drop(cols, axis =1)
    final_pmi = final_pmi.drop_duplicates().reset_index(drop= True)
    print('Final PMI Sheet Created...')
    
    # file_dir = os.path.dirname(os.path.abspath(__file__)

    print('Writing Sheets into WSWS Template...')
    fs = FileSystemStorage()
    f = StringIO('')
    fs.save('prt_sheet.csv', f)
    file_path = fs.path('prt_sheet.csv')
    file_url = fs.url('prt_sheet.csv')
    fs.save('pmi_sheet.csv', f)
    file_path1 = fs.path('pmi_sheet.csv')
    file_url1 = fs.url('pmi_sheet.csv')
    final_prt.to_csv(file_path, index=False)
    final_pmi.to_csv(file_path1, index=False)
    df_s1 = pd.read_csv(file_path, index_col='Unique #')
    df_s2 = pd.read_csv(file_path1, index_col='DMS# (To-Be)')


  
    filename = file4

    def append_df_to_excel(filename, df, sheet_name, startrow,
                           truncate_sheet=False, 
                           **to_excel_kwargs):
        
        

        # ignore [engine] parameter if it was passed
        if 'engine' in to_excel_kwargs:
            to_excel_kwargs.pop('engine')

        writer = pd.ExcelWriter(filename, engine='openpyxl')

        # Python 2.x: define [FileNotFoundError] exception if it doesn't exist 
        try:
            FileNotFoundError
        except NameError:
            FileNotFoundError = IOError


        try:
            # try to open an existing workbook
            writer.book = load_workbook(filename)

            # get the last row in the existing Excel sheet
            # if it was not specified explicitly
            if startrow is None and sheet_name in writer.book.sheetnames:
                startrow = writer.book[sheet_name].max_row

            # truncate sheet
            if truncate_sheet and sheet_name in writer.book.sheetnames:
                # index of [sheet_name] sheet
                idx = writer.book.sheetnames.index(sheet_name)
                # remove [sheet_name]
                writer.book.remove(writer.book.worksheets[idx])
                # create an empty sheet [sheet_name] using old index
                writer.book.create_sheet(sheet_name, idx)

            # copy existing sheets
            writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
        except FileNotFoundError:
            # file does not exist yet, we will create it
            pass

        if startrow is None:
            startrow = 0

        # write out the new sheet
        df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

        # save the workbook
        writer.save()


    # ### Call function and Write to Template

    # In[59]:


    append_df_to_excel(filename, df_s1, sheet_name='PRT Activity Content', startrow=6,
                           truncate_sheet=False)

    append_df_to_excel(filename, df_s2, sheet_name='PMI Generation Info', startrow=1,
                           truncate_sheet=False)
    print('All Tasks Completed.. WSWS Ready for Download...')
    return [ ]