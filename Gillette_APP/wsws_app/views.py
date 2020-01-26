import os
import warnings
from openpyxl import load_workbook
warnings.filterwarnings('ignore')
from django.contrib.staticfiles.templatetags.staticfiles import static
from django.core.files.storage import FileSystemStorage
from django.shortcuts import render
from django.views.generic import TemplateView

from wsws.settings import BASE_DIR
from .utils import read_file
import openpyxl

class FileView(TemplateView):
    template_name = 'index.html'

    def post(self, request, *args, **kwargs):
        files = request.FILES
        fs = FileSystemStorage()
        if files.get('wsws'):
            filename = '{}'.format(files.get('wsws').name)
            file_4 = fs.save(filename, files.get('wsws'))
            file_4_path = fs.path(file_4)
        else:
            #file_4 = static('wsws_file/wsws_sheet.xlsx') #get the file name
            
            path = 'E:\\user_jeff\Jupyter_Notebooks\Project Gillette\Gillette_APP\static\wsws_file\wsws_sheet.xlsx'
            #wb = static('wsws_file/wsws_sheet.xlsx') #get the file name
            #wb = openpyxl.load_workbook('wsws_sheet.xlsx')
            wb = openpyxl.load_workbook(path)
            # sheet_prt = wb['PRT Activity Content'] #get the sheet name
            # sheet_pmi = wb['PMI Generation Info']
            sheet_prt = wb['PRT Activity Content'] #get the sheet name
            sheet_pmi = wb['PMI Generation Info']

            # for a in sheet_prt['A7':'AU900000']: #you can set the range here 
            #    for cell in a:
            #      cell.value = None #set a value or null here


            # for a in sheet_pmi['A2':'U900000']: #you can set the range here 
            #    for cell in a:
            #      cell.value = None #set a value or null here

            wb.save('E:\\user_jeff\Jupyter_Notebooks\Project Gillette\Gillette_APP\static\wsws_file\wsws_sheet_master.xlsx')
            file_4 = static('wsws_file\wsws_sheet_master.xlsx')

            file_4_path = BASE_DIR + file_4
        output = []
        output += read_file(files.get('lmi'), files.get('cl'), files.get('ts'), file_4_path)
        output.append(file_4 if 'static' in file_4 else fs.url(file_4))
        output = map(lambda x: {"type": x.split('.')[1], "url": request.build_absolute_uri(x)}, output)
        return render(request, self.template_name, {"files": output})
